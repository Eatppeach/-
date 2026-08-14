"""敏感数据识别引擎 — 正则匹配与关键词检测"""

import re
import io
from database import get_enabled_rules, get_enabled_whitelist


class RecognitionEngine:
    """敏感数据识别引擎，支持多模式匹配"""

    def __init__(self):
        self._compiled_rules = []
        self._whitelist = []
        self._reload()

    def _reload(self):
        """从数据库重新加载规则并编译正则"""
        rules = get_enabled_rules()
        self._compiled_rules = []
        for rule in rules:
            if rule["rule_type"] == "regex":
                try:
                    compiled = re.compile(rule["pattern"], re.IGNORECASE)
                    self._compiled_rules.append({
                        "id": rule["id"],
                        "name": rule["name"],
                        "type": "regex",
                        "compiled": compiled,
                        "sensitivity_level": rule["sensitivity_level"],
                        "category": rule["category"],
                    })
                except re.error:
                    print(f"[WARN] 规则 '{rule['name']}' 正则编译失败，已跳过")
            elif rule["rule_type"] == "keyword":
                self._compiled_rules.append({
                    "id": rule["id"],
                    "name": rule["name"],
                    "type": "keyword",
                    "pattern": rule["pattern"],
                    "sensitivity_level": rule["sensitivity_level"],
                    "category": rule["category"],
                })

        self._whitelist = get_enabled_whitelist()

    def refresh(self):
        """热刷新规则（供管理API调用）"""
        self._reload()

    def _is_whitelisted(self, text, client_ip=None, user_id=None):
        """检查是否在白名单中"""
        for wl in self._whitelist:
            if wl["whitelist_type"] == "ip" and client_ip and wl["whitelist_value"] == client_ip:
                return True
            if wl["whitelist_type"] == "user" and user_id and wl["whitelist_value"] == user_id:
                return True
            if wl["whitelist_type"] == "text" and wl["whitelist_value"] in text:
                return True
        return False

    def scan(self, text, client_ip=None, user_id=None):
        """
        扫描文本中的敏感信息
        返回: {
            "has_sensitive": bool,
            "matches": [{"rule_name": ..., "matched_text": ..., "sensitivity_level": ..., "category": ..., "position": (start, end)}, ...],
            "match_count": int,
            "highest_level": str
        }
        """
        if not text:
            return {"has_sensitive": False, "matches": [], "match_count": 0, "highest_level": None}

        # 白名单检查
        if self._is_whitelisted(text, client_ip, user_id):
            return {"has_sensitive": False, "matches": [], "match_count": 0, "highest_level": None, "whitelisted": True}

        matches = []
        level_order = {"high": 3, "medium": 2, "low": 1}
        highest_level = None
        highest_level_val = 0

        for rule in self._compiled_rules:
            if rule["type"] == "regex":
                for m in rule["compiled"].finditer(text):
                    matched_text = m.group()
                    matches.append({
                        "rule_name": rule["name"],
                        "rule_id": rule["id"],
                        "matched_text": matched_text,
                        "sensitivity_level": rule["sensitivity_level"],
                        "category": rule["category"],
                        "position": (m.start(), m.end()),
                    })
                    lv = level_order.get(rule["sensitivity_level"], 0)
                    if lv > highest_level_val:
                        highest_level_val = lv
                        highest_level = rule["sensitivity_level"]
            elif rule["type"] == "keyword":
                try:
                    for m in re.finditer(rule["pattern"], text, re.IGNORECASE):
                        matched_text = m.group()
                        matches.append({
                            "rule_name": rule["name"],
                            "rule_id": rule["id"],
                            "matched_text": matched_text,
                            "sensitivity_level": rule["sensitivity_level"],
                            "category": rule["category"],
                            "position": (m.start(), m.end()),
                        })
                        lv = level_order.get(rule["sensitivity_level"], 0)
                        if lv > highest_level_val:
                            highest_level_val = lv
                            highest_level = rule["sensitivity_level"]
                except re.error:
                    pass

        return {
            "has_sensitive": len(matches) > 0,
            "matches": matches,
            "match_count": len(matches),
            "highest_level": highest_level,
        }

    def scan_file(self, file_bytes, filename):
        """扫描文件内容中的敏感信息"""
        text = self._extract_file_text(file_bytes, filename)
        return self.scan(text)

    def _extract_file_text(self, file_bytes, filename):
        """从文件中提取文本内容"""
        ext = filename.lower().split(".")[-1] if "." in filename else ""

        try:
            if ext == "txt":
                return file_bytes.decode("utf-8", errors="ignore")
            elif ext == "pdf":
                return self._extract_pdf_text(file_bytes)
            elif ext in ("docx", "doc"):
                return self._extract_docx_text(file_bytes)
            elif ext in ("xlsx", "xls"):
                return self._extract_xlsx_text(file_bytes)
            else:
                return file_bytes.decode("utf-8", errors="ignore")
        except Exception as e:
            print(f"[WARN] 文件解析失败: {filename}, 错误: {e}")
            return ""

    def _extract_pdf_text(self, file_bytes):
        try:
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text
        except ImportError:
            return ""
        except Exception:
            return ""

    def _extract_docx_text(self, file_bytes):
        try:
            from docx import Document
            doc = Document(io.BytesIO(file_bytes))
            return "\n".join([p.text for p in doc.paragraphs])
        except ImportError:
            return ""
        except Exception:
            return ""

    def _extract_xlsx_text(self, file_bytes):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(c) if c else "" for c in row])
                    if row_text.strip():
                        texts.append(row_text)
            wb.close()
            return "\n".join(texts)
        except ImportError:
            return ""
        except Exception:
            return ""


# 全局单例
recognition_engine = RecognitionEngine()